from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

for row in ws.iter_rows(min_row=2):
    # 영어점수 갖고오기
    if int(row[1].value) > 80:
        print(row[0].value, "번 학생은 영어천재")

for row in ws.iter_rows(max_row=1):
    for cell in row:
        if cell.value == "영어":  # 영어를
            cell.value = "컴퓨터"  # 컴퓨터로 바꿈

wb.save("sample_modified.xlsx")
