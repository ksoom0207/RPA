from random import *
from openpyxl import Workbook
wb = Workbook()

ws = wb.active
ws.title = "SOOMSheet"

ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3
ws["B1"] = 1
ws["B2"] = 2
ws["B3"] = 3

print(ws["A1"])  # A1의 셀 정보 출력
print(ws["A1"].value)  # 값출력
print(ws["E10"].value)  # 셀값이 없으면 NONE이라고 뜸

# 다른방법으로 접근
ws.cell(row=1, column=1)  # A1이랑 같음 row 1 column A
print(ws.cell(row=1, column=1).value)
print(ws.cell(row=1, column=2).value)  # ws[B1].value

ws.cell(row=3, column=4, value=10)  # == ws["C1"] = 10

# 반복문 대입
index = 1
for x in range(1, 20):
    for y in range(1, 20):  # 20x20개의 컬럼
       # ws.cell(row=x, column=y, value=randint(0, 100))
        ws.cell(row=x, column=y, value=index)
        index += 1  # 반복문 반대로 하면 열부터 데이터가 채워지겟징?

wb.save("sample2.xlsx")
