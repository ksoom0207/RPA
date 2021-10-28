from openpyxl import Workbook
wb = Workbook()  # 새 워크북 생성

ws = wb.create_sheet()  # 새로운 시트 기본이름으로 생성
ws.title = "MySheet"  # 이름변경
ws.sheet_properties.tabColor = "bb4422"  # RGB형태

ws1 = wb.create_sheet("YSheet")  # 주어진 이름으로 시트생성
# 위치지정
ws2 = wb.create_sheet("newsheet", 2)  # 2번째 인덱스에생성
# 시트 접근
new_ws = wb["newsheet"]  # Dict 형태로 시트에접근

# 시트이름 확인
print(wb.sheetnames)

# 시트내용 복사
new_ws["A1"] = "TEST"  # 셀의 값을 test로 입력
target = wb.copy_worksheet(new_ws)  # targer변수 설정해서 복사
target.title = "copied sheet"  # 이름바꾸기


wb.save("sample.xlsx")
wb.close()
