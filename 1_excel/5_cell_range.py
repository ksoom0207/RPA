from typing import Tuple
from openpyxl.utils.cell import coordinate_from_string
from openpyxl import Workbook
from random import *
wb = Workbook()
ws = wb.active

# 1 줄씩 데이터 넣기
ws.append(["번호", "영어", "수학"])  # A, B, C컬럼
# 데이터
for i in range(1, 11):
    ws.append([i, randint(0, 100), randint(0, 100)])

# 영어점수만 가져오기
# col_B = ws["B"]  # 영어컬럼나 가지고 오기
# # print(col_B) # <Cell 'Sheet'.B1>, <Cell 'Sheet'.B2> 이런식으로 출력
# for cell in col_B:
#     print(cell.value)

# 함께 가져올떄

# col_range = ws["B:C"]  # 컬럼함께 가지고오기

# for cols in col_range:
#     for cols2 in cols:
#         print(cols2.value)

# # Q. 2개를 갖고 왔슴에도 한줄만 출력하고 싶다면?

row_title = ws[1]  # 1번째 로우만 갖고오기

# for cell in row_title:
#     print(cell.value)

row_range = ws[2:6]  # 2번째 줄에서 6번째 줄까지 갖고오기
# for rows in row_range:  # 2-6낒; 한줄씩 갖고와서
#     for cell in rows:  # 그 줄에 있는걸 하나의 셀씩 끊어서 값출력
#         print(cell.value, end=" ")
#  #Q. end를 넣는이유는?  end를 넣어 준 이유는 해당 결괏값을 출력할 때 다음줄로 넘기지 않고 그 줄에 계속해서 출력하기 위해서이다.
#     print()

# 줄이많아서

row_ran = ws[2:ws.max_row]  # 2번째부터 마지막 줄 까지
for row in row_ran:
    for cell in row:
        # print(cell.value, end=" ")
        # print(cell.coordinate, end =" ") #CElL의 좌표정보 갖고옴
        xy = coordinate_from_string(cell.coordinate)  # ('A', 2) 튜플을 좌표상태처럼 끊어줌
        # print(xy, end=" ")
       # print(xy[0], end=" ")  # A B C A B C 이렇게 출력
        # print(xy[1], end=" ")  # 222 333 444 555 6 7 8 9 .. 이렇게 출력
       # print(xy[0:1], end=" ")  # ('A',) ('B',) ('C',)
        # print(xy[0:2], end=" ")  # ('A', 2) ('B', 2) ('C', 2)
        # 이거 안됨 ; print(xy.index(3)) #Q. 하나만 갖고 오는 방법은?
# 전체 rows 한줄씩
    # print(tuple(ws.rows)) #A1 B1 C1 이렇게 시작
# for row in tuple(ws.rows):
#     print(row[1].value)  # 1번째 셀 영어 주르륵 나옴(세로)

# for row in ws.iter_rows():
#     print(row[1].value) # 영어점수 갖고옴
# # 전체 한 열씩 가져와서
#     # print(tuple(ws.columns)) #A1 A2 이렇게 시작
# for column in tuple(ws.columns):
#     print(column[0].value) #번호 영어 수학 출력(가로)

# for colum in ws.iter_colw(): #전체 컬럼을 갖고옴
#     print(colum[1].value) # 영어점수 갖고옴

# iter_row장점?
# 범위를 지정해서 끊어가죠올수 있음 1번줄 부터 5번줄 까지 갖고옴
for row in ws.iter_rows(min_row=1, max_row=5, min_col=2, max_col=3):
    print(row[0].value, row[1].value)  # 수학점수 갖고옴 4번째 학생까지
    print(row)
# 출력형태
# 번호 영어 수학
# 1   2    3
    # 상하상하

for cols in ws.iter_cols(min_row=1, max_row=, min_col=1, max_col=3):  # col 3이상으로 하면 none나옴
   # print(cols[0].value, cols[1].value, cols[2].value)  # 수학점수 갖고옴 4번째 학생까지
    print(cols)
# 출력형태
# 번호 1 2
# 영어 38 23
# 수학 56 89
# 좌우좌우

# 데이터를 세로로 cols
# 데이터를 가로로 rows


# https://programtalk.com/python-examples/openpyxl.cell.coordinate_from_string/
wb.save("sample.xlsx")
# 자바스크립트 0.1 + 0.2 !== 0.3 맞음.. 부동소수점...
