from openpyxl import load_workbook
wb = load_workbook("sample2.xlsx")
ws = wb.active  # 활성화 된 Sheet

# for x in range(1, 20):
#     for y in range(1, 20):
#         print(ws.cell(row=x, column=y).value, end=" ")
#     print()

# cell 갯수모를때
for x in range(1, ws.max_row+1):
    for y in range(1, ws.max_column+1):
        print(ws.cell(row=x, column=y).value, end=",")
    print()
